"""
Script to compare the original and updated weather reports
"""

from openpyxl import load_workbook

def compare_reports():
    """
    Compare the original and updated weather reports
    """
    print("WEATHER REPORT COMPARISON")
    print("="*60)
    
    try:
        # Load both workbooks
        wb_original = load_workbook("city_weather_report.xlsx")
        wb_updated = load_workbook("city_weather_report_updated.xlsx")
        
        print("ORIGINAL REPORT:")
        print(f"  File: city_weather_report.xlsx")
        print(f"  Worksheets: {len(wb_original.worksheets)}")
        print(f"  Sheet names: {[ws.title for ws in wb_original.worksheets]}")
        
        ws_original = wb_original.active
        print(f"  Data rows: {ws_original.max_row - 1}")  # Subtract header row
        print(f"  Columns: {ws_original.max_column}")
        
        print("\nUPDATED REPORT:")
        print(f"  File: city_weather_report_updated.xlsx")
        print(f"  Worksheets: {len(wb_updated.worksheets)}")
        print(f"  Sheet names: {[ws.title for ws in wb_updated.worksheets]}")
        
        # Sheet 1 comparison
        ws1_updated = wb_updated["City Weather Report"]
        print(f"  Sheet 1 - Data rows: {ws1_updated.max_row - 1}")
        print(f"  Sheet 1 - Columns: {ws1_updated.max_column}")
        
        # Sheet 2 information
        ws2_updated = wb_updated["Cities by Region"]
        print(f"  Sheet 2 - Data rows: {ws2_updated.max_row - 1}")
        print(f"  Sheet 2 - Columns: {ws2_updated.max_column}")
        
        print("\n" + "="*60)
        print("KEY IMPROVEMENTS IN UPDATED VERSION:")
        print("="*60)
        
        # Count cities in each region
        regions = {}
        for row in ws1_updated.iter_rows(min_row=2, values_only=True):
            if row[1]:  # If region exists
                region = row[1]
                regions[region] = regions.get(region, 0) + 1
        
        print("1. ADDITIONAL CITIES:")
        print(f"   Original: 10 cities")
        print(f"   Updated: {ws1_updated.max_row - 1} cities")
        print(f"   Added: {ws1_updated.max_row - 1 - 10} new cities")
        
        print("\n2. REGIONAL BREAKDOWN:")
        for region, count in sorted(regions.items()):
            print(f"   {region}: {count} cities")
        
        print("\n3. NEW FEATURES:")
        print("   ✓ Second worksheet with regional grouping")
        print("   ✓ Blank rows between regions for readability")
        print("   ✓ Simplified regional view (City, Temperature, Region)")
        print("   ✓ Better organization for regional analysis")
        
        print("\n4. ENHANCED DATA ANALYSIS:")
        print("   ✓ Regional temperature averages")
        print("   ✓ Regional temperature ranges")
        print("   ✓ Easy comparison between regions")
        print("   ✓ Multiple views of the same data")
        
        print("\n" + "="*60)
        print("CLIENT BENEFITS:")
        print("="*60)
        print("• Quick regional analysis without searching through all data")
        print("• Clear visual separation between regions")
        print("• Focused view on key metrics (City, Temperature, Region)")
        print("• Professional presentation with multiple worksheet tabs")
        print("• Easy navigation between complete data and regional views")
        
    except FileNotFoundError as e:
        print(f"Error: {e}")
        print("Please ensure both report files exist in the current directory.")
    except Exception as e:
        print(f"Error comparing reports: {e}")

if __name__ == "__main__":
    compare_reports()

