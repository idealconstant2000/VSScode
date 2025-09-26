"""
Script to read and display the contents of the generated Excel report
"""

from openpyxl import load_workbook

def read_excel_report(filename="city_weather_report.xlsx"):
    """
    Read and display the contents of the Excel report
    """
    try:
        # Load the workbook
        wb = load_workbook(filename)
        ws = wb.active
        
        print(f"Reading Excel report: {filename}")
        print(f"Worksheet name: {ws.title}")
        print(f"Number of rows: {ws.max_row}")
        print(f"Number of columns: {ws.max_column}")
        print("\n" + "="*80)
        
        # Display all data
        for row in ws.iter_rows(values_only=True):
            if row[0]:  # Skip empty rows
                print(f"{row[0]:<15} {row[1]:<12} {row[2]:<12} {row[3]:<12} {row[4]:<12}")
        
        print("="*80)
        
        # Show some statistics
        print("\nSTATISTICS:")
        print("-" * 40)
        
        # Get data rows (skip header)
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        if data_rows:
            # Calculate statistics
            temperatures = [row[3] for row in data_rows if row[3]]
            humidities = [row[4] for row in data_rows if row[4]]
            populations = [row[2] for row in data_rows if row[2]]
            
            print(f"Average Temperature: {sum(temperatures)/len(temperatures):.1f}°F")
            print(f"Average Humidity: {sum(humidities)/len(humidities):.1f}%")
            print(f"Total Population: {sum(populations):,}")
            print(f"Average Population: {sum(populations)/len(populations):,.0f}")
            
            # Find hottest and coldest cities
            max_temp_idx = temperatures.index(max(temperatures))
            min_temp_idx = temperatures.index(min(temperatures))
            
            print(f"\nHottest City: {data_rows[max_temp_idx][0]} ({max(temperatures)}°F)")
            print(f"Coldest City: {data_rows[min_temp_idx][0]} ({min(temperatures)}°F)")
            
            # Find most and least humid cities
            max_humidity_idx = humidities.index(max(humidities))
            min_humidity_idx = humidities.index(min(humidities))
            
            print(f"Most Humid City: {data_rows[max_humidity_idx][0]} ({max(humidities)}%)")
            print(f"Least Humid City: {data_rows[min_humidity_idx][0]} ({min(humidities)}%)")
        
    except FileNotFoundError:
        print(f"Error: File '{filename}' not found.")
        print("Please run weather_analysis.py first to generate the report.")
    except Exception as e:
        print(f"Error reading Excel file: {e}")

if __name__ == "__main__":
    read_excel_report()

