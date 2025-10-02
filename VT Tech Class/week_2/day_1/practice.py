# ITP Week 2 Day 1 (In-Class) Practice

# A1. from the appropriate library, import only the Workbook
from openpyxl import Workbook

# A2. Before anything, we need a workbook to work with..
wb = Workbook()


# A3. We need to interact with a single worksheet.
ws = wb.active

# A4. assign the value of "First Name" to A1
ws['A1'] = "First Name"

# A5. assign the value of "Last Name" to B1
ws['B1'] = "Last Name"

# STOP HERE - RETURN TO LECTURE

# B1. For all of column A, starting at row 2 until row 10, make the cell values: "Gabriel" (attempt a loop)
first_names = ['Gabriel', 'John', 'Mary', 'Sophia', 'Emma', 'Olivia', 'Ava', 'Isabella', 'Mia'] 
for i in range(2, 11):
    ws[f'A{i}'] = first_names[i - 2]  # Adjust index since list starts at 0 and range starts at 2   




last_names = ['Rolley', 'Smith', 'Balenga', 'Issac', 'Cruise', 'Depp', 'Heard', 'Qiao', 'Biden']

# B2. Loop through a range from row 2 to 10 and assign the cell value to last names according to index in column B
# NOTE: PAY ATTENTION to the starting number of the range and how it differs from the starting index of the list
for i in range(2, 11):
    ws[f'B{i}'] = last_names[i - 2]  # Adjust index since list starts at 0 and range starts at 2    


# B3. Save the file
wb.save("wk2_day_1_practice.xlsx")