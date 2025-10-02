# ITP Week 3 Day 1 Practice

# import your required modules/methods

import openpyxl
from openpyxl import Workbook

# Create a new workbook
wb = Workbook()
ws = wb.active

# given the following items, using the methods we covered, write to openpyxl
headers = ["id", "name", "base_experience", "height", "order", "weight"]

for col_num, header in enumerate(headers, 1):  # start at column 1
    ws.cell(row=1, column=col_num, value=header)

# use an external counter with just a for loop (no function)
clefairy = {
    "id": 35,
    "name": "clefairy",
    "base_experience": 113,
    "height": 6,
    "order": 56,
    "weight": 75,
}

# external counter for column placement
col = 1
for key, value in clefairy.items():
    ws.cell(row=1, column=col, value=key)    # write header
    ws.cell(row=2, column=col, value=value)  # write value
    col += 1

# create a function that takes in a pokemon and writes its data
def write_pokemon(pokemon, start_row):
    col = 1
    for key, value in pokemon.items():
        ws.cell(row=start_row, column=col, value=value)
        col += 1


# create a function that takes in a pokemon
weedle = {
    "id": 13,
    "name": "weedle",
    "base_experience": 39,
    "height": 3,
    "order": 17,
    "weight": 32
}

# call the function with weedle!
write_pokemon(weedle, start_row=3)

# wb.save('./spreadsheets/practice.xlsx')

wb.save("./week_3/spreadsheets/practice_wk3_day1.xlsx")
print("Excel file 'inventory.xlsx' has been created!")