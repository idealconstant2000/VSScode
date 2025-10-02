import requests
import json

url = "https://jsonplaceholder.typicode.com/posts"
response = requests.get(url)

# Convert JSON string into Python objects (list of dictionaries)
data = json.loads(response.text)

#print(data[:1])  # Just look at the first post to understand the structure

#print(data)  # Look at the structure

from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Posts"

# Write headers (keys from the first post)
for col, key in enumerate(data[0].keys(), 1):
    ws.cell(row=1, column=col, value=key)

# write API data into excel sheet
for row, post in enumerate(data[:5], 2):  # start at row 2 (below headers)
    for col, value in enumerate(post.values(), 1):
        ws.cell(row=row, column=col, value=str(value))


wb.save("./week_3/spreadsheets/practice_wk3_day3.xlsx")
print("Excel file 'inventory.xlsx' has been created!")