"""
Script to read and display the contents of the updated Excel report with both sheets
"""

from openpyxl import load_workbook

def read_excel_report(filename="city_weather_report_updated.xlsx"):
    """
    Read and display the contents of both sheets in the updated Excel report
    """
    try:
        # Load the workbook
        wb = load_workbook(filename)
        
        print(f"Reading updated Excel report: {filename}")
        print(f"Number of worksheets: {len(wb.worksheets)}")
        print(f"Worksheet names: {[ws.title for ws in wb.worksheets]}")
        
        # Read Sheet 1: City Weather Report
        print("\n" + "="*80)
        print("SHEET 1: CITY WEATHER REPORT")
        print("="*80)
        
        ws1 = wb["City Weather Report"]
        print(f"Number of rows: {ws1.max_row}")
        print(f"Number of columns: {ws1.max_column}")
        print()
        
        # Display all data from Sheet 1
        for row in ws1.iter_rows(values_only=True):
            if row[0]:  # Skip empty rows
                print(f"{row[0]:<15} {row[1]:<12} {row[2]:<12} {row[3]:<12} {row[4]:<12}")
        
        # Read Sheet 2: Cities by Region
        print("\n" + "="*80)
        print("SHEET 2: CITIES BY REGION")
        print("="*80)
        
        ws2 = wb["Cities by Region"]
        print(f"Number of rows: {ws2.max_row}")
        print(f"Number of columns: {ws2.max_column}")
        print()
        
        # Display all data from Sheet 2
        for row in ws2.iter_rows(values_only=True):
            if row[0]:  # Skip empty rows
                print(f"{row[0]:<15} {row[1]:<12} {row[2]:<12}")
            else:
                print()  # Print blank line for spacing between regions
        
        # Show statistics for both sheets
        print("\n" + "="*80)
        print("STATISTICS")
        print("="*80)
        
        # Statistics from Sheet 1
        data_rows_sheet1 = list(ws1.iter_rows(min_row=2, values_only=True))
        if data_rows_sheet1:
            temperatures = [row[3] for row in data_rows_sheet1 if row[3]]
            humidities = [row[4] for row in data_rows_sheet1 if row[4]]
            populations = [row[2] for row in data_rows_sheet1 if row[2]]
            
            print("OVERALL STATISTICS (from Sheet 1):")
            print(f"  Average Temperature: {sum(temperatures)/len(temperatures):.1f}°F")
            print(f"  Average Humidity: {sum(humidities)/len(humidities):.1f}%")
            print(f"  Total Population: {sum(populations):,}")
            print(f"  Average Population: {sum(populations)/len(populations):,.0f}")
            
            # Find hottest and coldest cities
            max_temp_idx = temperatures.index(max(temperatures))
            min_temp_idx = temperatures.index(min(temperatures))
            
            print(f"  Hottest City: {data_rows_sheet1[max_temp_idx][0]} ({max(temperatures)}°F)")
            print(f"  Coldest City: {data_rows_sheet1[min_temp_idx][0]} ({min(temperatures)}°F)")
        
        # Regional statistics from Sheet 2
        print("\nREGIONAL BREAKDOWN (from Sheet 2):")
        current_region = None
        region_temps = {}
        region_cities = {}
        
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if row[0] and row[2]:  # If it's a data row (not empty)
                region = row[2]
                city = row[0]
                temp = row[1]
                
                if region not in region_temps:
                    region_temps[region] = []
                    region_cities[region] = []
                
                region_temps[region].append(temp)
                region_cities[region].append(city)
        
        for region in sorted(region_temps.keys()):
            temps = region_temps[region]
            cities = region_cities[region]
            print(f"  {region}:")
            print(f"    Cities: {len(cities)} ({', '.join(cities)})")
            print(f"    Average Temperature: {sum(temps)/len(temps):.1f}°F")
            print(f"    Temperature Range: {min(temps)}°F - {max(temps)}°F")
        
    except FileNotFoundError:
        print(f"Error: File '{filename}' not found.")
        print("Please run weather_analysis_updated.py first to generate the report.")
    except Exception as e:
        print(f"Error reading Excel file: {e}")

if __name__ == "__main__":
    read_excel_report()

