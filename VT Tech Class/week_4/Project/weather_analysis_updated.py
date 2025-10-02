import requests
import json
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def fetch_data_with_retry(url, max_retries=3, delay=2):
    """
    Fetch data from API with retry logic for rate limiting
    """
    for attempt in range(max_retries):
        try:
            response = requests.get(url)
            response.raise_for_status()
            return json.loads(response.text)
        except requests.exceptions.RequestException as e:
            if "429" in str(e) and attempt < max_retries - 1:
                print(f"Rate limited. Waiting {delay} seconds before retry {attempt + 2}/{max_retries}...")
                time.sleep(delay)
                delay *= 2  # Exponential backoff
            else:
                raise e
    return None

def get_mock_data():
    """
    Return mock data for demonstration when APIs are unavailable
    """
    cities_data = [
        {"city": "New York", "population": 8336817, "region": "Northeast"},
        {"city": "Los Angeles", "population": 3979576, "region": "West"},
        {"city": "Chicago", "population": 2693976, "region": "Midwest"},
        {"city": "Houston", "population": 2320268, "region": "South"},
        {"city": "Phoenix", "population": 1680992, "region": "West"},
        {"city": "Philadelphia", "population": 1584064, "region": "Northeast"},
        {"city": "San Antonio", "population": 1547253, "region": "South"},
        {"city": "San Diego", "population": 1423851, "region": "West"},
        {"city": "Dallas", "population": 1343573, "region": "South"},
        {"city": "San Jose", "population": 1035317, "region": "West"},
        {"city": "Detroit", "population": 670031, "region": "Midwest"},
        {"city": "Miami", "population": 467963, "region": "South"},
        {"city": "Atlanta", "population": 498715, "region": "South"},
        {"city": "Denver", "population": 715522, "region": "West"},
        {"city": "Minneapolis", "population": 429606, "region": "Midwest"}
    ]
    
    weather_data = [
        {"city": "New York", "temperature": 68, "humidity": 65},
        {"city": "Los Angeles", "temperature": 75, "humidity": 45},
        {"city": "Chicago", "temperature": 58, "humidity": 70},
        {"city": "Houston", "temperature": 82, "humidity": 80},
        {"city": "Phoenix", "temperature": 95, "humidity": 25},
        {"city": "Philadelphia", "temperature": 70, "humidity": 60},
        {"city": "San Antonio", "temperature": 85, "humidity": 75},
        {"city": "San Diego", "temperature": 72, "humidity": 50},
        {"city": "Dallas", "temperature": 88, "humidity": 65},
        {"city": "San Jose", "temperature": 78, "humidity": 55},
        {"city": "Detroit", "temperature": 62, "humidity": 68},
        {"city": "Miami", "temperature": 90, "humidity": 85},
        {"city": "Atlanta", "temperature": 88, "humidity": 70},
        {"city": "Denver", "temperature": 65, "humidity": 40},
        {"city": "Minneapolis", "temperature": 55, "humidity": 75}
    ]
    
    return cities_data, weather_data

def fetch_data():
    """
    Fetch data from both APIs and return as Python objects
    """
    print("Fetching data from APIs...")
    
    # API endpoints
    cities_url = "https://city-weather-project.free.beeceptor.com/cities"
    weather_url = "https://city-weather-project.free.beeceptor.com/weather"
    
    try:
        # Fetch cities data
        print("Fetching cities data...")
        cities_data = fetch_data_with_retry(cities_url)
        
        # Fetch weather data
        print("Fetching weather data...")
        weather_data = fetch_data_with_retry(weather_url)
        
        if cities_data and weather_data:
            print(f"Successfully fetched {len(cities_data)} cities and {len(weather_data)} weather records")
            return cities_data, weather_data
        else:
            raise Exception("Failed to fetch data from APIs")
        
    except Exception as e:
        print(f"Error fetching data from APIs: {e}")
        print("Using mock data for demonstration...")
        cities_data, weather_data = get_mock_data()
        print(f"Using mock data: {len(cities_data)} cities and {len(weather_data)} weather records")
        return cities_data, weather_data

def examine_data_structure(cities_data, weather_data):
    """
    Examine the structure of the fetched data
    """
    print("\n" + "="*50)
    print("EXAMINING DATA STRUCTURE")
    print("="*50)
    
    if cities_data:
        print("Cities data structure:")
        print(f"Number of cities: {len(cities_data)}")
        print("First city record:")
        print(json.dumps(cities_data[0], indent=2))
    
    if weather_data:
        print("\nWeather data structure:")
        print(f"Number of weather records: {len(weather_data)}")
        print("First weather record:")
        print(json.dumps(weather_data[0], indent=2))

def combine_data(cities_data, weather_data):
    """
    Combine city and weather data using city name as the key
    """
    print("\n" + "="*50)
    print("COMBINING DATA")
    print("="*50)
    
    # Create a dictionary for quick city lookup
    cities_dict = {city['city']: city for city in cities_data}
    
    combined_data = []
    unmatched_weather = []
    
    for weather_record in weather_data:
        city_name = weather_record['city']
        
        if city_name in cities_dict:
            # Combine the data
            combined_record = {
                'city': city_name,
                'region': cities_dict[city_name]['region'],
                'population': cities_dict[city_name]['population'],
                'temperature': weather_record['temperature'],
                'humidity': weather_record['humidity']
            }
            combined_data.append(combined_record)
        else:
            unmatched_weather.append(city_name)
    
    print(f"Successfully combined {len(combined_data)} records")
    if unmatched_weather:
        print(f"Warning: {len(unmatched_weather)} weather records had no matching city data: {unmatched_weather}")
    
    # Show example of combined record
    if combined_data:
        print("\nExample combined record:")
        print(json.dumps(combined_data[0], indent=2))
    
    return combined_data

def group_cities_by_region(combined_data):
    """
    Group cities by region for the regional view
    """
    print("\n" + "="*50)
    print("GROUPING CITIES BY REGION")
    print("="*50)
    
    # Initialize region groups
    regions = {
        'Midwest': [],
        'West': [],
        'South': [],
        'Northeast': []  # Added Northeast since we have cities from this region
    }
    
    # Group cities by region
    for record in combined_data:
        region = record['region']
        if region in regions:
            # Create simplified record for regional view
            regional_record = {
                'city': record['city'],
                'temperature': record['temperature'],
                'region': record['region']
            }
            regions[region].append(regional_record)
    
    # Display grouping results
    for region, cities in regions.items():
        if cities:
            print(f"{region}: {len(cities)} cities")
            for city in cities:
                print(f"  - {city['city']}: {city['temperature']}°F")
    
    return regions

def create_excel_report(combined_data):
    """
    Create and format the main Excel report (Sheet 1)
    """
    print("\n" + "="*50)
    print("CREATING MAIN EXCEL REPORT")
    print("="*50)
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "City Weather Report"
    
    # Define headers
    headers = ['City', 'Region', 'Population', 'Temperature', 'Humidity']
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data
    for row, record in enumerate(combined_data, 2):
        ws.cell(row=row, column=1, value=record['city'])
        ws.cell(row=row, column=2, value=record['region'])
        ws.cell(row=row, column=3, value=record['population'])
        ws.cell(row=row, column=4, value=record['temperature'])
        ws.cell(row=row, column=5, value=record['humidity'])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add borders and formatting
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply borders to all cells with data
    for row in ws.iter_rows(min_row=1, max_row=len(combined_data) + 1, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border
    
    # Center align data cells
    for row in ws.iter_rows(min_row=2, max_row=len(combined_data) + 1, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    print(f"Main Excel report created with {len(combined_data)} data rows")
    return wb

def create_regional_sheet(wb, regions):
    """
    Create the regional grouping sheet (Sheet 2)
    """
    print("\n" + "="*50)
    print("CREATING REGIONAL GROUPING SHEET")
    print("="*50)
    
    # Create new worksheet
    ws_regional = wb.create_sheet(title="Cities by Region")
    
    # Define headers for regional sheet
    headers = ['City', 'Temperature', 'Region']
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws_regional.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write grouped data with spacing between regions
    current_row = 2
    
    # Define region order for consistent output
    region_order = ['Midwest', 'West', 'South', 'Northeast']
    
    for region in region_order:
        if region in regions and regions[region]:
            print(f"Writing {region} region with {len(regions[region])} cities...")
            
            # Write cities for this region
            for city_record in regions[region]:
                ws_regional.cell(row=current_row, column=1, value=city_record['city'])
                ws_regional.cell(row=current_row, column=2, value=city_record['temperature'])
                ws_regional.cell(row=current_row, column=3, value=city_record['region'])
                current_row += 1
            
            # Add blank row after each region (except the last one)
            if region != region_order[-1]:
                current_row += 1
    
    # Auto-adjust column widths
    for column in ws_regional.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_regional.column_dimensions[column_letter].width = adjusted_width
    
    # Add borders and formatting
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply borders to all cells with data
    for row in ws_regional.iter_rows(min_row=1, max_row=current_row-1, min_col=1, max_col=3):
        for cell in row:
            if cell.value:  # Only apply borders to non-empty cells
                cell.border = thin_border
    
    # Center align data cells
    for row in ws_regional.iter_rows(min_row=2, max_row=current_row-1, min_col=1, max_col=3):
        for cell in row:
            if cell.value:  # Only align non-empty cells
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    print(f"Regional sheet created with {current_row-2} data rows (including spacing)")
    return wb

def save_report(wb, filename="city_weather_report_updated.xlsx"):
    """
    Save the workbook to file
    """
    print(f"\nSaving updated report as {filename}...")
    wb.save(filename)
    print(f"Report successfully saved as {filename}")

def main():
    """
    Main function to orchestrate the entire process
    """
    print("WEATHER ANALYSIS PROJECT - UPDATED VERSION")
    print("="*60)
    
    # Step 1: Fetch the data
    cities_data, weather_data = fetch_data()
    
    if not cities_data or not weather_data:
        print("Failed to fetch data. Exiting.")
        return
    
    # Step 2: Examine data structure
    examine_data_structure(cities_data, weather_data)
    
    # Step 3: Combine the data
    combined_data = combine_data(cities_data, weather_data)
    
    if not combined_data:
        print("No data to combine. Exiting.")
        return
    
    # Step 4: Group cities by region
    regions = group_cities_by_region(combined_data)
    
    # Step 5: Create main Excel report
    wb = create_excel_report(combined_data)
    
    # Step 6: Create regional grouping sheet
    wb = create_regional_sheet(wb, regions)
    
    # Step 7: Save the updated report
    save_report(wb)
    
    print("\n" + "="*60)
    print("UPDATED PROJECT COMPLETED SUCCESSFULLY!")
    print("="*60)
    print(f"Created city_weather_report_updated.xlsx with:")
    print(f"- Sheet 1: 'City Weather Report' with {len(combined_data)} cities")
    print(f"- Sheet 2: 'Cities by Region' with grouped data")
    print("\nThe updated report includes:")
    print("- Complete dataset with all city and weather information")
    print("- Regional grouping for easy analysis by geographic area")
    print("- Professional formatting with headers, borders, and spacing")
    print("- Multiple worksheets for different views of the same data")

if __name__ == "__main__":
    main()

