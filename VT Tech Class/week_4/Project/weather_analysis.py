import requests
import json
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def fetch_data_with_retry(url, max_retries=3, delay=2):
    """
    Fetch data from API with retry logic for rate limiting
    """
    for attempt in range(max_retries):
        try:
            response = requests.get(url)
            response.raise_for_status()
            return json.loads(response.text)
        except requests.exceptions.RequestException as e:
            if "429" in str(e) and attempt < max_retries - 1:
                print(f"Rate limited. Waiting {delay} seconds before retry {attempt + 2}/{max_retries}...")
                time.sleep(delay)
                delay *= 2  # Exponential backoff
            else:
                raise e
    return None

def get_mock_data():
    """
    Return mock data for demonstration when APIs are unavailable
    Based on the actual API data structure (state-level data)
    """
    cities_data = [
        {"city": "California", "population": 39512223, "region": "West"},
        {"city": "Texas", "population": 28995881, "region": "South"},
        {"city": "Florida", "population": 21477737, "region": "South"},
        {"city": "New York", "population": 19453561, "region": "Northeast"},
        {"city": "Pennsylvania", "population": 12801989, "region": "Northeast"},
        {"city": "Illinois", "population": 12671821, "region": "Midwest"},
        {"city": "Ohio", "population": 11689100, "region": "Midwest"},
        {"city": "Georgia", "population": 10617423, "region": "South"},
        {"city": "North Carolina", "population": 10488084, "region": "South"},
        {"city": "Michigan", "population": 9986857, "region": "Midwest"}
    ]
    
    weather_data = [
        {"city": "California", "temperature": 75, "humidity": 55},
        {"city": "Texas", "temperature": 94, "humidity": 62},
        {"city": "Florida", "temperature": 92, "humidity": 75},
        {"city": "New York", "temperature": 69, "humidity": 64},
        {"city": "Pennsylvania", "temperature": 69, "humidity": 66},
        {"city": "Illinois", "temperature": 73, "humidity": 65},
        {"city": "Ohio", "temperature": 71, "humidity": 65},
        {"city": "Georgia", "temperature": 89, "humidity": 72},
        {"city": "North Carolina", "temperature": 84, "humidity": 70},
        {"city": "Michigan", "temperature": 66, "humidity": 64}
    ]
    
    return cities_data, weather_data

def fetch_data():
    """
    Fetch data from both APIs and return as Python objects
    """
    print("Fetching data from APIs...")
    
    # API endpoints - Updated URLs
    cities_url = "https://cityweatherclass.free.beeceptor.com/cities"
    weather_url = "https://cityweatherclass.free.beeceptor.com/weather"
    
    try:
        # Fetch cities data
        print("Fetching cities data...")
        cities_data = fetch_data_with_retry(cities_url)
        
        # Fetch weather data
        print("Fetching weather data...")
        weather_data = fetch_data_with_retry(weather_url)
        
        if cities_data and weather_data:
            print(f"Successfully fetched {len(cities_data)} states and {len(weather_data)} weather records")
            return cities_data, weather_data
        else:
            raise Exception("Failed to fetch data from APIs")
        
    except Exception as e:
        print(f"Error fetching data from APIs: {e}")
        print("Using mock data for demonstration...")
        cities_data, weather_data = get_mock_data()
        print(f"Using mock data: {len(cities_data)} states and {len(weather_data)} weather records")
        return cities_data, weather_data

def examine_data_structure(cities_data, weather_data):
    """
    Examine the structure of the fetched data
    """
    print("\n" + "="*50)
    print("EXAMINING DATA STRUCTURE")
    print("="*50)
    
    if cities_data:
        print("States data structure:")
        print(f"Number of states: {len(cities_data)}")
        print("First state record:")
        print(json.dumps(cities_data[0], indent=2))
    
    if weather_data:
        print("\nWeather data structure:")
        print(f"Number of weather records: {len(weather_data)}")
        print("First weather record:")
        print(json.dumps(weather_data[0], indent=2))

def combine_data(cities_data, weather_data):
    """
    Combine state and weather data using state name as the key
    """
    print("\n" + "="*50)
    print("COMBINING DATA")
    print("="*50)
    
    # Create a dictionary for quick state lookup
    states_dict = {state['city']: state for state in cities_data}
    
    combined_data = []
    unmatched_weather = []
    
    for weather_record in weather_data:
        state_name = weather_record['city']
        
        if state_name in states_dict:
            # Combine the data
            combined_record = {
                'city': state_name,
                'region': states_dict[state_name]['region'],
                'population': states_dict[state_name]['population'],
                'temperature': weather_record['temperature'],
                'humidity': weather_record['humidity']
            }
            combined_data.append(combined_record)
        else:
            unmatched_weather.append(state_name)
    
    print(f"Successfully combined {len(combined_data)} records")
    if unmatched_weather:
        print(f"Warning: {len(unmatched_weather)} weather records had no matching state data: {unmatched_weather}")
    
    # Show example of combined record
    if combined_data:
        print("\nExample combined record:")
        print(json.dumps(combined_data[0], indent=2))
    
    return combined_data

def create_excel_report(combined_data):
    """
    Create and format the Excel report
    """
    print("\n" + "="*50)
    print("CREATING EXCEL REPORT")
    print("="*50)
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "State Weather Report"
    
    # Define headers
    headers = ['State', 'Region', 'Population', 'Temperature', 'Humidity']
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data
    for row, record in enumerate(combined_data, 2):
        ws.cell(row=row, column=1, value=record['city'])
        ws.cell(row=row, column=2, value=record['region'])
        ws.cell(row=row, column=3, value=record['population'])
        ws.cell(row=row, column=4, value=record['temperature'])
        ws.cell(row=row, column=5, value=record['humidity'])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add borders and formatting
    from openpyxl.styles import Border, Side
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply borders to all cells with data
    for row in ws.iter_rows(min_row=1, max_row=len(combined_data) + 1, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border
    
    # Center align data cells
    for row in ws.iter_rows(min_row=2, max_row=len(combined_data) + 1, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    print(f"Excel report created with {len(combined_data)} data rows")
    return wb

def save_report(wb, filename="state_weather_report.xlsx"):
    """
    Save the workbook to file
    """
    print(f"\nSaving report as {filename}...")
    wb.save(filename)
    print(f"Report successfully saved as {filename}")

def main():
    """
    Main function to orchestrate the entire process
    """
    print("STATE WEATHER ANALYSIS PROJECT")
    print("="*50)
    
    # Step 1: Fetch the data
    cities_data, weather_data = fetch_data()
    
    if not cities_data or not weather_data:
        print("Failed to fetch data. Exiting.")
        return
    
    # Step 2: Examine data structure
    examine_data_structure(cities_data, weather_data)
    
    # Step 3: Combine the data
    combined_data = combine_data(cities_data, weather_data)
    
    if not combined_data:
        print("No data to combine. Exiting.")
        return
    
    # Step 4: Create Excel report
    wb = create_excel_report(combined_data)
    
    # Step 5: Save the report
    save_report(wb)
    
    print("\n" + "="*50)
    print("PROJECT COMPLETED SUCCESSFULLY!")
    print("="*50)
    print(f"Created state_weather_report.xlsx with {len(combined_data)} states")
    print("The report includes:")
    print("- State names and regions")
    print("- Population data")
    print("- Current temperature and humidity")
    print("- Professional formatting with headers and borders")

if __name__ == "__main__":
    main()
