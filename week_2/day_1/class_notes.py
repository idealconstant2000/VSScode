import openpyxl
from openpyxl import Workbook

wb = Workbook()
print(type(wb))   # <class 'openpyxl.workbook.workbook.Workbook'>

ws = wb.active # Get the default worksheet
ws.title = "First Sheet" # Rename the default sheet

ws1 = wb.create_sheet("Rugrats")        # adds at the end
ws2 = wb.create_sheet("Hey Arnold", 0)  # adds at the beginning
















wb.save("test_spreadsheet.xlsx ")