# ITP Week 2 Day 2 Exercise

# import the appropriate method from the correct module
from openpyxl import load_workbook 
# Import the workbook that you updated in today's practice from
# "./spreadsheets/inventory.xlsx"
wb = load_workbook("./week_2/spreadsheets/inventory.xlsx")
# access and store the appropriate worksheet to the variable 'ws'
ws = wb['CURRENT_MONTH_INVENTORY']

# Define a function called add_order_amount that takes in a single parameter called 'row'

    # IF the quantity is less or equal to than the threshold,

        # than calculate the order_amount (max_amount - quantity) 
        # assign the value to that row, column 6
# TIP: create variables for quantity, threshold, max_amount that retrieves the values first for cleanliness

def add_order_amount(row):
    max_amount = ws.cell(row=row, column=3).value
    threshold = ws.cell(row=row, column=4).value
    quantity = ws.cell(row=row, column=5).value

    if quantity <= threshold:
        order_amount = max_amount - quantity
        ws.cell(row=row, column=6).value = order_amount
    else:
        ws.cell(row=row, column=6).value = 0

# Perform a for..in loop through the range(2, len(inventory.rows))
#   - call the function add_order_amount() passing in the number of the range
for row in range(2, ws.max_row + 1):
    add_order_amount(row)

# save the latest changes   
wb.save("./week_2/spreadsheets/inventory_day2_exercise.xlsx")