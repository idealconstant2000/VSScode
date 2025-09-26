import time
import json
import requests
from openpyxl import Workbook

# API endpoints
character_url = "https://rickandmortyapi.com/api/character"
location_url  = "https://rickandmortyapi.com/api/location"
episode_url   = "https://rickandmortyapi.com/api/episode"

# Workbook + sheets
wb = Workbook()
ws_characters = wb.active
ws_characters.title = "Rick and Morty Characters"
ws_locations = wb.create_sheet("Rick and Morty Locations")
ws_episodes  = wb.create_sheet("Rick and Morty Episodes")

session = requests.Session()
session.headers.update({"Accept": "application/json"})
lookup_cache = {}

def safe_get_json(url, retries=3, timeout=15):
    """GET a URL and return JSON or None (never raises JSONDecodeError). Retries on 429/5xx/timeouts."""
    backoff = 0.8
    for attempt in range(retries):
        try:
            resp = session.get(url, timeout=timeout)
            # Retry politely on rate limiting
            if resp.status_code == 429:
                retry_after = resp.headers.get("Retry-After")
                wait = float(retry_after) if retry_after else backoff
                time.sleep(wait)
                backoff *= 2
                continue

            # Raise for other bad HTTP statuses
            resp.raise_for_status()

            # Ensure it's JSON
            ctype = resp.headers.get("Content-Type", "")
            if "application/json" not in ctype:
                return None

            return resp.json()
        except requests.RequestException:
            # network/timeout/HTTP errors -> backoff and retry
            time.sleep(backoff)
            backoff *= 2
        except ValueError:
            # JSON decode problem
            time.sleep(backoff)
            backoff *= 2
    return None

def get_name_from_url(url):
    """Return the 'name' for a Rick & Morty API resource URL, with caching & safety."""
    if not url or not isinstance(url, str) or not url.startswith("http"):
        return url
    if url in lookup_cache:
        return lookup_cache[url]

    data = safe_get_json(url)
    if isinstance(data, dict) and "name" in data:
        name = data["name"]
    else:
        # Couldn’t decode/resolve; fall back to the URL (or last path segment if you prefer)
        # name = url.rsplit("/", 1)[-1]
        name = url

    lookup_cache[url] = name
    return name

def replace_urls_with_names(value):
    """Replace a single URL or list of URLs with names; leave other values untouched."""
    # Single URL string
    if isinstance(value, str) and value.startswith("http"):
        return get_name_from_url(value)
    # List of URLs
    if isinstance(value, list):
        # Only treat as URL list if elements are strings that look like URLs
        if value and all(isinstance(v, str) and v.startswith("http") for v in value):
            return [get_name_from_url(v) for v in value]
        return value
    # Dicts (e.g., origin/location objects) -> replace embedded 'url' if present
    if isinstance(value, dict):
        v = dict(value)  # shallow copy
        if "url" in v and isinstance(v["url"], str) and v["url"].startswith("http"):
            v["name"] = get_name_from_url(v["url"])  # augment with name
            # optionally drop the url after resolving:
            # v.pop("url", None)
        return v
    return value

def fetch_and_write_all_pages(start_url, ws):
    """Follow info['next'] to exhaust all pages, resolving URLs to names safely."""
    url = start_url
    row = 2
    headers_written = False
    headers = []

    while url:
        data = safe_get_json(url)
        if not data:
            # Couldn’t load this page; stop politely rather than crashing
            break

        results = data.get("results", [])
        if not results:
            break

        if not headers_written:
            headers = list(results[0].keys())
            for col, key in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=key)
            headers_written = True

        for item in results:
            for col, key in enumerate(headers, start=1):
                value = item.get(key)

                # Only attempt URL-to-name replacement if a URL exists
                # (your NOTE about using if-statements)
                if value is not None:
                    value = replace_urls_with_names(value)

                # Flatten complex types so Excel can store them
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)

                ws.cell(row=row, column=col, value=value)
            row += 1

        url = (data.get("info") or {}).get("next")

# Run all three sheets
fetch_and_write_all_pages(character_url, ws_characters)
fetch_and_write_all_pages(location_url, ws_locations)
fetch_and_write_all_pages(episode_url, ws_episodes)

# Save workbook
wb.save("./week_3/spreadsheets/exercise-nightmare.xlsx")
print("Wrote exercise-nightmare.xlsx with names resolved instead of URLs")
