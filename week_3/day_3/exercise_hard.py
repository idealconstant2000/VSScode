import json
import requests
from openpyxl import Workbook

# API endpoints
character_url = "https://rickandmortyapi.com/api/character"
location_url = "https://rickandmortyapi.com/api/location"
episode_url = "https://rickandmortyapi.com/api/episode"

# Workbook + sheets
wb = Workbook()
ws_characters = wb.active
ws_characters.title = "Rick and Morty Characters"
ws_locations = wb.create_sheet("Rick and Morty Locations")
ws_episodes = wb.create_sheet("Rick and Morty Episodes")

def fetch_and_write_all_pages(start_url, ws):
    """Follow info['next'] to exhaust all pages, writing headers once."""
    url = start_url
    row = 2
    headers_written = False
    headers = []

    total_written = 0
    session = requests.Session()

    while url:
        resp = session.get(url, timeout=20)
        resp.raise_for_status()
        data = resp.json()

        results = data.get("results", [])
        if not results:
            break

        # Write headers exactly once from the first page's item
        if not headers_written:
            headers = list(results[0].keys())
            for col, key in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=key)
            headers_written = True

        # Write all rows on this page
        for item in results:
            for col, key in enumerate(headers, start=1):
                value = item.get(key)
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)
                ws.cell(row=row, column=col, value=value)
            row += 1
            total_written += 1

        # Hop to the next page via info.next
        url = data.get("info", {}).get("next")

    return total_written

# Populate all three worksheets (fetch EVERYTHING)
count_chars = fetch_and_write_all_pages(character_url, ws_characters)
count_locs = fetch_and_write_all_pages(location_url, ws_locations)
count_eps  = fetch_and_write_all_pages(episode_url, ws_episodes)

# Save workbook
wb.save("./week_3/spreadsheets/exercise-hard.xlsx")
print(f"Wrote exercise-hard.xlsx with {count_chars} characters, {count_locs} locations, {count_eps} episodes.")
