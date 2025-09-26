# ITP Week 3 Day 3 Exercise

# RICK AND MORTY API DOCS: https://rickandmortyapi.com/documentation

# we want to make a copy of the Rick and Morty database (which is provided through the api)


# EASY MODE

# import the appropriate modules (you have 3)
import json
import requests
from openpyxl import Workbook

character_url = "https://rickandmortyapi.com/api/character"

# Set up a workbook and worksheet titled "Rick and Morty Characters"
wb = Workbook()
ws_characters = wb.active
ws_characters.title = "Rick and Morty Characters"

row = 2  # start writing data on row 2 (row 1 is for headers)
headers_written = False
headers = []

url = character_url
while url:
    resp = requests.get(url)
    resp.raise_for_status()
    data = resp.json()  # same as json.loads(resp.text)

    # Write headers once, based on the first character's keys
    if not headers_written:
        headers = list(data["results"][0].keys())
        for col, key in enumerate(headers, start=1):
            ws_characters.cell(row=1, column=col, value=key)
        headers_written = True

    # Write character rows
    for character in data["results"]:
        for col, key in enumerate(headers, start=1):
            value = character.get(key)
            # Convert dicts/lists to strings so openpyxl can write them
            if isinstance(value, (dict, list)):
                value = json.dumps(value)  # prettier than str()
            ws_characters.cell(row=row, column=col, value=value)
        row += 1

    # Follow pagination
    url = data.get("info", {}).get("next")

# Save the workbook
wb.save("./week_3/spreadsheets/exercise.xlsx")
print("Wrote exercise.xlsx")





# NOTE: due to the headers, the rows need to be offset by one!

# MEDIUM MODE

# create 2 new worksheets for "Rick and Morty Locations" and "Rick and Morty Episodes"

# create 2 new variables for episode_url and location_url (retrieve it from the docs!)
# "https://rickandmortyapi.com/api/location"
# "https://rickandmortyapi.com/api/episode"

# populate the new worksheets appropriately with all of the data!

# NOTE: don't forget your headers!

# HARD MODE
# Can you decipher the INFO key of the data to use "next" url to continuously pull data?
# Currently, we are only pulling 20 items per api pull!
# WE WANT EVERYTHING. (contact instructors for office hours if stuck!)

# NIGHTMARE
# The inner information for characters, locations, and episodes, references one another through urls
# ie. for episode 28, it lists all the character but by their url
# can you use the URLs to make a subsequent call inside your for loops
# to replace the url with just the appropriate names?
# NOTE: need to make use of if statements to see if url exists or not
# (contact instructors for office hours if stuck!)

