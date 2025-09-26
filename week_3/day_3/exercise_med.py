import json
import requests
from openpyxl import Workbook

# API endpoints
character_url = "https://rickandmortyapi.com/api/character"
location_url = "https://rickandmortyapi.com/api/location"
episode_url = "https://rickandmortyapi.com/api/episode"

# Workbook + sheets
wb = Workbook()
ws_characters = wb.active
ws_characters.title = "Rick and Morty Characters"
ws_locations = wb.create_sheet("Rick and Morty Locations")
ws_episodes = wb.create_sheet("Rick and Morty Episodes")

def fetch_and_write_all_pages(url, ws):
    """Fetch paginated results from an endpoint and write to the given worksheet."""
    row = 2                      # row 1 reserved for headers
    headers_written = False
    headers = []

    while url:
        resp = requests.get(url)
        resp.raise_for_status()
        data = resp.json()

        results = data.get("results", [])
        if not results:
            break

        # Write headers based on the first item (once)
        if not headers_written:
            headers = list(results[0].keys())
            for col, key in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=key)
            headers_written = True

        # Write rows
        for item in results:
            for col, key in enumerate(headers, start=1):
                value = item.get(key)
                # Convert nested dicts/lists to JSON strings
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)
                ws.cell(row=row, column=col, value=value)
            row += 1

        # pagination
        url = data.get("info", {}).get("next")

# Populate all three worksheets
fetch_and_write_all_pages(character_url, ws_characters)
fetch_and_write_all_pages(location_url, ws_locations)
fetch_and_write_all_pages(episode_url, ws_episodes)

# Save workbook
wb.save("./week_3/spreadsheets/exercise-med.xlsx")
print("Wrote exercise-med.xlsx")
