import openpyxl
from openpyxl import Workbook

# Create a new workbook
wb = Workbook()

# Select the active worksheet
ws = wb.active

headers = ["Product ID", "Name", "Reorder Threshold", "Inventory", "Max Amount", "Description"]

for col_num, header in enumerate(headers, 1):  # start at column 1
    ws.cell(row=1, column=col_num, value=header)

item_oreo = {
    "product_id": 2323,
    "name": "oreo",
    "reorder_threshold": 300,
    "inventory": 743,
    "max_amount": 1000,
    "description": "yummy yummy"
}

# ws.cell(row=1, column=1, value=item_oreo['product_id'])
# ws.cell(row=1, column=2, value=item_oreo['name'])
# ws.cell(row=1, column=3, value=item_oreo['reorder_threshold'])
# ws.cell(row=1, column=4, value=item_oreo['inventory'])
# ws.cell(row=1, column=5, value=item_oreo['max_amount'])

# def dict_range_to_row(item, row_number):
#     for col_number in range(1, len(item.keys()) + 1):
#         ws.cell(row=row_number, column=col_number, value=list(item.values())[col_number - 1])

# def dict_keys_to_row(item, row_number):
#     col_count = 1
#     for key in item:
#         ws.cell(row=row_number, column=col_count, value=item[key])
#         col_count += 1

# def items_dict_to_row(item, row_number):
#     col_count = 1
#     for key, value in item.items():
#         ws.cell(row=row_number, column=col_count, value=value)
#         col_count += 1

def dict_enumerate_to_row(item, row_number):
    for col_count, value in enumerate(item.values(), 1):  # start at 1 for Excel
        ws.cell(row=row_number, column=col_count, value=value)

item_coke = {
    "product_id": 6545,
    "name": "coke",
    "reorder_threshold": 100,
    "inventory": 101,
    "max_amount": 500,
    "discription": "classic"
}

item_pepsi = {
    "product_id": 3456,
    "name": "pepsi",
    "reorder_threshold": 50,
    "inventory": 137,
    "max_amount": 200,
    "description": "the choice of a new generation"
}

inventory_list = [item_oreo, item_coke, item_pepsi]

for row_count, each_dict in enumerate(inventory_list, 2): # start rows at 2 (headers in row 1)
    dict_enumerate_to_row(each_dict, row_count)

# Save the workbook
wb.save("./week_3/spreadsheets/inventory.xlsx")
print("Excel file 'inventory.xlsx' has been created!")

 


